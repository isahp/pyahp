#This is the class that has all of the data for a pw comparison

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import json

#The dictionary of symbolic values
Symbolic_Vote_To_Number = {
    ">": -1,
    ">>": -2,
    "<": -3,
    "<<": -4,
    "E":1,
    "e":1
}


Symbolic_Inverse = {
    -1: -3,
    -2: -4,
    -3:-1,
    -4:-2,
    1:1
}

def addMatrixPlace(nparr):
    shape = nparr.shape
    size = shape[0]
    if size == 0:
        return(np.identity(1))
    rval = np.c_[nparr, np.zeros(size)]
    rval = np.r_[rval, [np.zeros(size+1)]]
    rval[size, size] = 1.0
    return(rval)

def largest_eigen(npmat, error = 1e-8):
    shape = npmat.shape
    if len(shape) != 2:
        raise NameError("Wrong shape")
    if shape[0] != shape[1]:
        raise NameError("Not square")
    size = shape[0]
    vec = np.ones(size)
    close = False
    while not close:
        nextVec = np.matmul(npmat, vec)
        nextVec = nextVec/max(nextVec)
        diff = max(abs(nextVec - vec))
        close = (diff < error)
        vec = nextVec
    return(nextVec)


def symbolic_inverse(aval):
    if aval in Symbolic_Inverse:
        return Symbolic_Inverse[aval]
    elif aval == 0:
        return 0
    else:
        return 1.0/aval

def setReflexive(arr, row, col, val):
    arr[row, col] = val
    if val == 0:
        arr[col,row] = 0
    else:
        arr[col,row] = symbolic_inverse(val)


def symbolic_vote_value(theVote):
    if theVote in Symbolic_Vote_To_Number:
        return Symbolic_Vote_To_Number[theVote]
    else:
        return(theVote)

def pw_model_from_excel(excel_file):
    rval = PwModel()
    rval.loadExcel(excel_file)
    return(rval)

class PwModel:
    def __init__(self):
        self.clear()

    def clear(self):
        self.alts = []
        self.users = []
        self.pws = {}
        self.nalts = 0
        self.groups = {}
        self.better = 3
        self.much_better = 9

    def addAlt(self, alt, raiseError=True):
        if alt in self.alts:
            if raiseError:
                raise NameError("Alt already existed")
            else:
                return()
        #Add the alt to the alt names list
        self.alts.append(alt)
        self.nalts+=1
        #Add the alt to each of the matrices
        for uname in self.users:
            self.pws[uname] = addMatrixPlace(self.pws[uname])


    def addUser(self, uname):
        if uname in self.users:
            raise NameError("User already existed")
        self.users.append(uname)
        mat = np.identity(self.nalts)
        self.pws[uname] = mat

    def loadExcel(self, excelFile):
        #Clear out old data
        self.clear()
        wb = load_workbook(excelFile)
        for sheet in wb.sheetnames:
            if not sheet == "info":
                #Add the user for the sheet
                self.addUser(sheet)
                self.parseExcelSheet(sheet, wb[sheet])
                print("Got a sheet "+sheet)
            else:
                self.parseExcelInfo(wb[sheet])

    def setVote(self, uname, alt1, alt2, vote):
        alt1Index = self.alts.index(alt1)
        alt2Index = self.alts.index(alt2)
        setReflexive(self.pws[uname], alt1Index, alt2Index, vote)

    def parseExcelInfo(self, wbsheet):
        allgroups = {}
        uNames = wbsheet.columns[0]
        uNames = [cell.value for cell in uNames ]
        for colIndex in range(1, len(wbsheet.columns)):
            col = wbsheet.columns[colIndex]
            groups = {}
            colName = col[0].value
            allgroups[colName] = groups
            for rowIndex in range(1, len(col)):
                value = col[rowIndex].value
                usersWithValue = groups.get(value, None)
                if usersWithValue == None:
                    usersWithValue = []
                    groups[value] = usersWithValue
                usersWithValue.append(uNames[rowIndex])
        self.groups = allgroups


    def parseExcelSheet(self, uname, wbsheet):
        for row in wbsheet.iter_rows():
            maxcol = len(row)
            if maxcol == 3 and row[2].value != None:
                #We have a pairwise comparison
                alt1 = row[0].value
                alt2 = row[2].value
                symvote = row[1].value
                vote = symbolic_vote_value(symvote)
                #Add the alts to be safe
                self.addAlt(alt1, raiseError=False)
                self.addAlt(alt2, raiseError=False)
                self.setVote(uname, alt1, alt2, vote)

    def getVoteValue(self, aVote):
        if aVote < 0:
            if aVote == -1.:
                #This is better
                return(self.better)
            elif aVote == -2.:
                #This is much better
                return(self.much_better)
            elif aVote == -3.:
                #This is opposite better
                return(1.0/self.better)
            elif aVote == -4.:
                return(1.0/self.much_better)
            else:
                return(1.0)
        else:
            return(aVote)

    def getValueMatrix(self, user:str):
        mat = np.copy(self.pws[user])
        size = mat.shape[0]
        for row in range(size):
            for col in range(size):
                mat[row, col] = self.getVoteValue(mat[row,col])
        return(mat)

    def getGroupValueMatrix(self, users):
        matrices = [self.getValueMatrix(user) for user in users]
        rval = np.identity(self.nalts)
        for row in range(self.nalts):
            for col in range(self.nalts):
                rval[row,col]=1.0
                count = 0
                for v in range(len(matrices)):
                    val = matrices[v][row,col]
                    if not val == 0.0:
                        count += 1
                        rval[row,col] *= val
                if count == 0:
                    rval[row, col] = 0.0
                else:
                    rval[row, col] = pow(rval[row, col], 1.0/count)
        return(rval)

    def getUserPriorities(self):
        rval = {user: largest_eigen(self.getValueMatrix(user)).tolist() for user in self.users}
        return(rval)

    def getGroupPriorities(self):
        rval = {group:{subgroup: largest_eigen(self.getGroupValueMatrix(self.groups[group][subgroup])).tolist() for subgroup in self.groups[group]} for group in self.groups}
        return(rval)

    def getAllCalcs(self):
        rval = {
            "alts": self.alts,
            "users": self.users,
            "userScores": self.getUserPriorities(),
            "groups": self.groups,
            "groupScores": self.getGroupPriorities()
        }
        return(rval)


if __name__ == '__main__':
    md = PwModel()
    md.addUser("Bill")
    md.addUser("Bill2")
    md.addAlt("alt1")
    md.addAlt("alt2")
    print(md.pws)
    md = pw_model_from_excel('Jobs.xlsx')
    print(md.pws)
    print(md.groups)
    print(md.getVoteValue(-3))
    mat = md.getValueMatrix('Percy')
    eig = largest_eigen(mat)
    print(mat)
    print(eig)
    userInfos = md.getUserPriorities()
    print(userInfos)
    print(md.getGroupValueMatrix(md.groups['Gender']['M']))
    print(md.getGroupPriorities())
    rvalJson = json.dumps(md.getAllCalcs())